import pandas as pd 
import numpy as np 
from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
import win32com.client
from pywintypes import com_error
import webbrowser

root = Tk()
root.title("GORKHA COMPANY TRAINING CENTER")
root.configure(background="#b3c1f2")
root.option_add( "*font", "Verdana  18" )  #setting default font and fontsize
root.iconbitmap("bgphoto_icon.ico")
#making frame
insert_frame= LabelFrame(root, padx=2,pady=5, bg="#8290bf")
view_frame= LabelFrame(root, bg="#455959")
image_frame=LabelFrame(insert_frame,bg="#DF0101",height=350,width=350)
#packing frame items 
insert_frame.grid( row=1, column= 0 ,padx=0, pady=1,sticky="w")
view_frame.grid( row=0, column= 0,padx=5, pady=5)
image_frame.grid(row=2,column=3,padx=1,pady=1,rowspan=5,columnspan=2,sticky="w")

#bg image
bgimg= PhotoImage(file= r"photo.png")
Label(image_frame, image=bgimg).place(relwidth=1,relheight=1)

#edit certain cell
def edit(name, col,val):
    data = pd.read_excel("demo.xlsx")
    wb = load_workbook("demo.xlsx")
    ws = wb.active
    c=1
    for colu in data.columns:
        if colu == col:
            break
        c+=1
    r=0
    for row in ws.rows:
        if (row[0].value.lower() == search_data.get().lower()) or (row[1].value.lower()== search_data.get().lower()):
            break
        r+=1
    ws.cell(row=r,column=c, value=val)
    wb.save(filename = 'demo.xlsx')
def add():
    data = pd.read_excel("demo.xlsx")
    wb = load_workbook("demo.xlsx")
    ws = wb.active
    b=0
    for row in ws.rows:
        if (row[0].value.lower() == rollnumber_data.get().lower()):
            b = 1
            break

    if (not name_data.get) or (not age_data.get()) or (not address_data.get()) or (not contact_data.get()) or (not amount_data.get()) or (not rollnumber_data.get()):
        pop("Empty Field","All field are necessary")
    elif b ==1:
        pop("Invalid", "Same roll number have been useed.")
    else:
        df = pd.DataFrame({'Roll No':[rollnumber_data.get()],
            'Name': [name_data.get()],
            "Father's Name":[fathers_name_data.get()],
            "GrandFather's Name":[grandfathers_name_data.get()],
            'Age': [age_data.get()],
            'Address':[address_data.get()],
            'Contact':[contact_data.get()],
            #'Amount Paid':[amount_data.get()],
            #'Month':[amount_data.get()]
            })
        writer = pd.ExcelWriter('demo.xlsx', engine='openpyxl')
# try to open an existing workbook
        writer.book = load_workbook('demo.xlsx')
# copy existing sheets
        writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
# read existing file
        reader = pd.read_excel(r'demo.xlsx')
# write out the new sheet
        df.to_excel(writer,index=False,header=False,startrow=len(reader)+1)
        writer.close()
        edit(name_data.get(),month.get(), amount_data.get())

        pop("Data enetred","Entered value sucessfully")
def pop(data1,data2):
     messagebox.showinfo(data1,data2)
#writer.close()
def delete(val):
    df = pd.read_excel('demo.xlsx', index_col=[0])
    df = df.drop([val], axis=0)
    df.to_excel("demo.xlsx")
    #print("Deleted sucessfully")
def search():
    #try:
    from PIL import ImageTk, Image
    top=Toplevel()
    top.title("View recruits info")
    wb = load_workbook("demo.xlsx")
    #image for deatil
    imagef1= LabelFrame(top,padx=200,pady=50,bg="#b3c1f2")
    imagef1.grid(row=0,column=0,sticky="ew")
    valueframe= LabelFrame(top,padx=200,pady=50,background="#b3c1f2")
    valueframe.grid(row=1,column=0)
    data = pd.read_excel("demo.xlsx")
    ws = wb.active
    r = 1
    d=0  
    import os
    desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 
    desktop = desktop + "\\student\\" +search_data.get()+".png"
    print(desktop)
    print(type(desktop))
    data1="Details about :" + search_data.get()
    Label(imagef1,text=data1).grid(row=0,column=0,sticky="ew",columnspan=3)
    colm=1
    for colu in data.columns:
        if colu == "Total":
            break
        colm+=1
    print("Column = "+ str(colm))
    r=0
    i=1
    for row in ws.rows:
        c = 1
        if (row[0].value == "Roll No"):
            for cell in row:
                val=Label(valueframe,text=cell.value).grid(row=c,column=r)
                c+=1
                print(i)
                if i == 7:
                    print('break')
                    break
                i+=1
            r+=1
            d1=ws.cell(row=r,column=colm)
            d1=d1.value
            val=Label(valueframe,text=d1).grid(row=c,column=r-1)
            print (d1)
        i=1
        if (row[0].value.lower() == search_data.get().lower()) or (row[1].value.lower() == search_data.get().lower()):
            for cell in row:
                val=Label(valueframe,text=cell.value).grid(row=c,column=r)
                c+=1 
                d=1
                if i == 7:
                    break
                i+=1
            r+=1
            d1=ws.cell(row=r,column=colm)
            d1=d1.value
            val=Label(valueframe,text=d1).grid(row=c,column=r-1)

        #bg image
    #bgimg= PhotoImage(file= desktop)
    #
    img = Image.open(desktop)
    img = img.resize((250,250), Image.ANTIALIAS) 
    img = ImageTk.PhotoImage(img)
    panel =  Label(valueframe, image=img).grid(row=0,column=4,sticky="w",rowspan=7).place(relwidth=1,relheight=1)
    #panel.image = img

    if d==0:
        Label(top,text="Sorry No data available for that name.").grid(row=r,column=c)
    '''except:
        pop("Unable to find","Unable to retrieve data.")'''
def open(value):
    top=Toplevel()
    top.title("View recruits info")
    Label(top,text=value).pack()
    print(value)
def view():
    
# Path to original excel file
    WB_PATH = r'C:\Users\Sandeep\Desktop\FoodBillingSystem_PYTHON\FoodBillingSystem_Py\demo.xlsx'
# PDF path when saving
    PATH_TO_PDF = r'C:\Users\Sandeep\Desktop\FoodBillingSystem_PYTHON\FoodBillingSystem_Py\demo6.pdf'


    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False
    
    try:
        print('Start conversion to PDF')

    # Open
        wb = excel.Workbooks.Open(WB_PATH)

    # Specify the sheet you want to save by index. 1 is the first (leftmost) sheet.
        ws_index_list = [1,2]
        wb.WorkSheets(ws_index_list).Select()

    # Save
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except:
        print('failed.')
    else:
        webbrowser.open_new(r'demo6.pdf')
        print('Succeeded.')
    finally:
        
        wb.Close()
        excel.Quit()
        

def update():
    pass
def add_photo():
    from tkinter import filedialog
    import os
    try:
        if not rollnumber_data.get():
            pop("Invalid","Enter roll number before selecting photo")
        else:
            rep = filedialog.askopenfilenames(
                parent=root,
                initialdir='/',
                initialfile='tmp',
                filetypes=[
                    ("All files", "*"),
                    ("PNG", "*.png"),
                    ("JPEG", "*.jpg"),
                    ("JPG","*.jpg")])
            rep1 = ''.join(rep)
    except IndexError:
        pop("Filed","Failed to load photo.") 
    addp(str(rep1))
def addp(name):
    try:
        data = pd.read_excel("demo.xlsx")
        wb = load_workbook("demo.xlsx")
        ws = wb.active
        b=0
        for row in ws.rows:
            if (row[0].value.lower() == rollnumber_data.get().lower()):
                b = 1
                break
        if (b ==1):
            pop("Unable to add photo","Student data with that roll number exists.")
        else:

            import os
            import shutil
            desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') 
            desktop = desktop + "\\student\\"
            if not os.path.isdir(desktop):
                os.mkdir(desktop)
            desktop=desktop+str(rollnumber_data.get())+".png"
    #os.rename(name, desktop)
            shutil.copyfile(name, desktop)
    #os.replace(name, desktop)
    except:
        pop("Failed","Failed to move photo")

name_data=StringVar()
fathers_name_data=StringVar()
grandfathers_name_data=StringVar()
age_data=StringVar()
address_data=StringVar()
contact_data=StringVar()
search_data=StringVar()
amount_data=StringVar()
month = StringVar()
rollnumber_data=StringVar()
#root.resizable(height=false,width =false)
#buttons
add_photo = Button(insert_frame,text="Add Photo",bg="#e4eded",fg="#063332",command=add_photo)
insert = Button(insert_frame, text="INSERT NEW RECRUITS",bg="#e4eded", fg="#063332",width=25, command=add)
view = Button(view_frame, text="VIEW RECRUITS INFO", bg="#e4eded", fg="#063332",command=view)
update = Button(view_frame, text="UPDATE RECRUITS INFO", bg="#e4eded", fg="#063332",command=update)
search = Button(view_frame, text="SEARCH", bg="#e4eded", fg="#063332",command=search)
#only for search
search_e = Entry(view_frame, textvariable=search_data)
search_e.grid(row=0, column=5)
add_photo.grid(row=8,column=0)
insert.grid( row=8, column= 2)
view.grid( row=0, column= 2)
update.grid( row=0, column= 3)
search.grid( row=0, column= 4)


#insert labels 
#root.wm_attributes('-transparentcolor','#8290bf')
rollnumber = Label(insert_frame,text="Roll NO:", padx=2,pady=2,bg="#8290bf",fg="white")
name = Label(insert_frame, text="Fullname :", padx=20, pady=20, bg="#8290bf",fg="white")
fathers_name = Label(insert_frame, text="Father's name :", padx=20, pady=20, bg="#8290bf",fg="white")
grandfathers_name = Label(insert_frame, text="Grandfather's name :", padx=20, pady=20, bg="#8290bf",fg="white")
age = Label(insert_frame, text="Age :", padx=20, pady=20, bg="#8290bf",fg="white")
address = Label(insert_frame, text="Address :", padx=20, pady=20, bg="#8290bf",fg="white")
contact = Label(insert_frame, text="Contact :", padx=20, pady=20, bg="#8290bf",fg="white")
amount = Label(insert_frame, text="Amount Paid :", padx=20, pady=20, bg="#8290bf",fg="white")

#entry
rollnumber_e = Entry(insert_frame,textvariable=rollnumber_data,width=12)
name_e = Entry(insert_frame, textvariable=name_data, width=25)
fathers_name_e = Entry(insert_frame, textvariable=fathers_name_data, width=25)
grandfathers_name_e = Entry(insert_frame, textvariable=grandfathers_name_data,width=25)
age_e = Entry(insert_frame, textvariable=age_data,width=25)
address_e = Entry(insert_frame, textvariable=address_data,width=25)
contact_e = Entry(insert_frame, textvariable=contact_data,width=25)
amount_e = Entry(insert_frame, textvariable=amount_data,width=12)

rollnumber.grid(row=1,column=3,sticky="w")
rollnumber_e.grid(row=1,column=4,sticky="w")
name.grid(row=1, column=0)
name_e.grid(row=1, column=2, sticky="ew")
fathers_name.grid(row=2, column=0)
fathers_name_e.grid(row=2, column=2, sticky="ew")
grandfathers_name.grid(row=3, column=0)
grandfathers_name_e.grid(row=3, column=2, sticky="ew")
age.grid(row=4, column=0)
age_e.grid(row=4, column=2, sticky="ew")
address.grid(row=5, column=0)
address_e.grid(row=5, column=2, sticky="ew")
contact.grid(row=6, column=0)
contact_e.grid(row=6, column=2, sticky="ew")
amount.grid(row=7, column=0)
amount_e.grid(row=7, column=2 , sticky = "w")
list1= ['Baisakh','Jestha','Ashad','Shrawan','Bhadra','Ashwin','Kartik','Mangsir','Poush','Magh','Falgun','Chaitra']
droplist = OptionMenu(insert_frame,month,*list1)
droplist.config(width = 10)
month.set('Baisakh')
droplist.grid(row = 7,column =2, sticky= "e")



root.mainloop()